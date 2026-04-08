"""
Re-aggregation Script for NL2 Profit and Loss Extractor.

Rebuilds the Master_Data sheet from individual company verification sheets.
"""

import sys
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import click
import openpyxl
from rich.console import Console

sys.path.insert(0, str(Path(__file__).parent))

from extractor.models import NL2Extract, NL2Data
from config.row_registry import NL2_ROW_ORDER
from validation.checks import run_validations, write_validation_report, build_validation_summary_table
from output.excel_writer import save_workbook

console = Console()
logger = logging.getLogger(__name__)


def parse_sheet_to_extract(ws) -> Optional[NL2Extract]:
    """Parse a company verification sheet back into a NL2Extract object."""
    try:
        meta_cell = ws.cell(row=2, column=1).value
        if not meta_cell or "Quarter:" not in meta_cell:
            return None

        parts = {p.split(":")[0].strip(): p.split(":")[1].strip() for p in meta_cell.split("|")}
        quarter = parts.get("Quarter", "Qx")
        year = parts.get("Year", "xxxxxx")
        source_file = parts.get("Source", "Unknown.pdf")

        title = ws.cell(row=1, column=1).value
        company_name = title.replace("VERIFICATION SHEET:", "").strip()

        from config.company_registry import COMPANY_DISPLAY_NAMES
        company_key = "unknown"
        for k, v in COMPANY_DISPLAY_NAMES.items():
            if v == company_name:
                company_key = k
                break

        extract = NL2Extract(
            source_file=source_file,
            company_key=company_key,
            company_name=company_name,
            form_type="NL2",
            quarter=quarter,
            year=year,
            data=NL2Data(),
        )

        # header row is row 5 (start_row=4 + 1), data starts row 6
        header_row = 5
        col_map = {"cy_qtr": 2, "cy_ytd": 3, "py_qtr": 4, "py_ytd": 5}

        for i, pl_key in enumerate(NL2_ROW_ORDER):
            ws_row = header_row + 1 + i
            pl_data = {}
            for period, col in col_map.items():
                val = ws.cell(row=ws_row, column=col).value
                try:
                    pl_data[period] = float(val) if val is not None else None
                except (ValueError, TypeError):
                    pl_data[period] = None
            extract.data.data[pl_key] = pl_data

        return extract
    except Exception as e:
        logger.error(f"Failed to parse sheet {ws.title}: {e}")
        return None


@click.command()
@click.option("--workbook", "-w", required=True,
              help="Path to the Excel master workbook", type=click.Path(exists=True))
@click.option("--backup/--no-backup", default=True,
              help="Create a timestamped backup before overwriting")
def reaggregate(workbook, backup):
    """Rebuild Master_Data sheet from individual company sheets."""
    console.print(f"[bold blue]Re-aggregating workbook:[/bold blue] {workbook}")

    if backup:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = Path(workbook).with_name(f"{Path(workbook).stem}_backup_{timestamp}.xlsx")
        shutil.copy2(workbook, backup_path)
        console.print(f"[dim]Backup created: {backup_path.name}[/dim]")

    try:
        wb = openpyxl.load_workbook(workbook)
        extractions = []

        for sheet_name in wb.sheetnames:
            if sheet_name in ["Master_Data", "_meta", "Validation_Summary", "Validation_Detail"]:
                continue
            ws = wb[sheet_name]
            extract = parse_sheet_to_extract(ws)
            if extract:
                extractions.append(extract)
                console.print(f"  [green]\u2713[/green] Parsed {sheet_name}")
            else:
                console.print(f"  [red]\u2717[/red] Skipped {sheet_name}")

        if not extractions:
            console.print("[bold red]No valid company sheets found.[/bold red]")
            return

        console.print(f"\n[bold blue]Running validations on {len(extractions)} companies...[/bold blue]")
        val_results = run_validations(extractions)
        report_path = Path(workbook).parent / "validation_report_reagg.csv"
        write_validation_report(val_results, str(report_path))
        save_workbook(extractions, workbook,
                      stats={"files_processed": len(extractions), "files_succeeded": len(extractions)})

        console.print(f"\n[bold green]Re-aggregation successful![/bold green]")
        console.print(build_validation_summary_table(val_results))

    except Exception as e:
        console.print(f"[bold red]Critical error:[/bold red] {e}")
        sys.exit(1)


if __name__ == "__main__":
    reaggregate()
