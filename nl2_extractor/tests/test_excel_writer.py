"""
test_excel_writer.py -- NL-2 Excel output tests.
"""

import os
import pytest
from openpyxl import load_workbook
from output.excel_writer import save_workbook
from extractor.models import NL2Extract, NL2Data
from config.settings import MASTER_COLUMNS


def _make_extract(pl_key: str = "profit_before_tax", cy_ytd: float = 211342.0):
    extract = NL2Extract(
        source_file="test.pdf",
        company_key="bajaj_allianz",
        company_name="Bajaj Allianz General Insurance Company Limited",
        form_type="NL2",
        quarter="Q3",
        year="202526",
        data=NL2Data(),
    )
    extract.data.data[pl_key] = {
        "cy_qtr": cy_ytd / 4,
        "cy_ytd": cy_ytd,
        "py_qtr": cy_ytd / 4 * 0.9,
        "py_ytd": cy_ytd * 0.9,
    }
    return extract


def test_save_workbook_creates_file(tmp_path):
    output_file = tmp_path / "nl2_test.xlsx"
    save_workbook([_make_extract()], str(output_file))
    assert os.path.exists(output_file)


def test_save_workbook_has_expected_sheets(tmp_path):
    output_file = tmp_path / "nl2_test.xlsx"
    save_workbook([_make_extract()], str(output_file))
    wb = load_workbook(output_file)
    assert "Master_Data" in wb.sheetnames
    assert "_meta" in wb.sheetnames
    assert "BajajAllianz_Q3_202526" in wb.sheetnames


def test_save_workbook_empty_list(tmp_path):
    output_file = tmp_path / "empty.xlsx"
    save_workbook([], str(output_file))
    assert os.path.exists(output_file)


def test_master_data_column_headers(tmp_path):
    """Master_Data row 1 must contain the NL2 column names."""
    output_file = tmp_path / "headers.xlsx"
    save_workbook([_make_extract()], str(output_file))
    wb = load_workbook(output_file)
    ws = wb["Master_Data"]
    header_row = [ws.cell(row=1, column=c).value for c in range(1, len(MASTER_COLUMNS) + 1)]
    assert "PL_PARTICULARS" in header_row
    assert "Hierarchy_Depth" in header_row
    assert "Year_Info" in header_row
    assert "Quarter_Info" in header_row
    assert "Value" in header_row
    # Wide-format columns must NOT be present
    assert "CY_Qtr" not in header_row
    assert "CY_YTD" not in header_row
    assert "PY_Qtr" not in header_row
    assert "PY_YTD" not in header_row
    # Other form columns must NOT be present
    assert "Gross_Commission" not in header_row
    assert "Total_Channel" not in header_row
    assert "LOB_PARTICULARS" not in header_row


def test_master_data_has_all_row_order_entries(tmp_path):
    """Master_Data must have 4 rows per NL2_ROW_ORDER data entry (depth != -1) — long format."""
    from config.row_registry import NL2_ROW_ORDER, NL2_ROW_DEPTH
    output_file = tmp_path / "rows.xlsx"
    extract = NL2Extract(
        source_file="test.pdf", company_key="bajaj_allianz",
        company_name="Bajaj", form_type="NL2", quarter="Q3", year="202526",
        data=NL2Data(),
    )
    for key in NL2_ROW_ORDER:
        extract.data.data[key] = {"cy_qtr": 1.0, "cy_ytd": 2.0, "py_qtr": 0.9, "py_ytd": 1.8}

    save_workbook([extract], str(output_file))
    wb = load_workbook(output_file)
    ws = wb["Master_Data"]
    row_count = ws.max_row - 1  # subtract header
    expected = 4 * sum(1 for k in NL2_ROW_ORDER if NL2_ROW_DEPTH.get(k, 1) != -1)
    assert row_count == expected, (
        f"Expected {expected} data rows (4 per item, excluding section headers), got {row_count}"
    )


def test_hierarchy_depth_column_written(tmp_path):
    """Hierarchy_Depth column must contain integer values."""
    from config.row_registry import NL2_ROW_DEPTH
    output_file = tmp_path / "depth.xlsx"
    save_workbook([_make_extract("total_a", 218314.0)], str(output_file))
    wb = load_workbook(output_file)
    ws = wb["Master_Data"]
    depth_col = MASTER_COLUMNS.index("Hierarchy_Depth") + 1
    depths = [ws.cell(row=r, column=depth_col).value for r in range(2, ws.max_row + 1)]
    # All depth values must be integers (0, 1, or 2)
    for d in depths:
        if d is not None:
            assert isinstance(d, int), f"Hierarchy_Depth must be int, got {type(d)}: {d}"
            assert d in (0, 1, 2), f"Unexpected depth value: {d}"


def test_cy_ytd_value_written(tmp_path):
    """CY YTD value must appear in the Value column of Master_Data."""
    output_file = tmp_path / "values.xlsx"
    save_workbook([_make_extract("profit_before_tax", 211342.0)], str(output_file))
    wb = load_workbook(output_file)
    ws = wb["Master_Data"]
    val_col = MASTER_COLUMNS.index("Value") + 1
    values = [ws.cell(row=r, column=val_col).value for r in range(2, ws.max_row + 1)]
    assert 211342.0 in values


def test_incremental_preserve_existing_rows(tmp_path):
    """Saving a new company must preserve rows from a previous company."""
    output_file = tmp_path / "incremental.xlsx"
    ext1 = _make_extract("profit_before_tax", 100.0)
    ext1.source_file = "company1.pdf"
    save_workbook([ext1], str(output_file))

    ext2 = NL2Extract(
        source_file="company2.pdf", company_key="icici_lombard",
        company_name="ICICI Lombard", form_type="NL2",
        quarter="Q3", year="202526", data=NL2Data(),
    )
    ext2.data.data["profit_before_tax"] = {"cy_ytd": 200.0}
    save_workbook([ext2], str(output_file))

    wb = load_workbook(output_file)
    ws = wb["Master_Data"]
    sf_col = MASTER_COLUMNS.index("Source_File") + 1
    source_files = {ws.cell(row=r, column=sf_col).value for r in range(2, ws.max_row + 1)}
    assert "company1.pdf" in source_files
    assert "company2.pdf" in source_files
