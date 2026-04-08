"""
Excel Writer for NL-2-B-PL Profit and Loss Account.

Output structure:
  - Master_Data sheet: one row per company per P&L line item
  - Per-company verification sheets: P&L rows x 4 period columns
  - _meta sheet: run metadata
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config.settings import (
    MASTER_COLUMNS,
    EXTRACTOR_VERSION,
    NUMBER_FORMAT,
    company_key_to_pascal,
)
from config.row_registry import NL2_ROW_ORDER, NL2_ROW_DISPLAY_NAMES, NL2_ROW_DEPTH
from config.pl_metadata import get_pl_particulars, get_grouped_pl
from config.company_metadata import get_metadata
from extractor.models import NL2Extract

logger = logging.getLogger(__name__)

# Style definitions
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_META_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_BOLD_FONT = Font(bold=True)
_SUMMARY_FILL = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
# Section header rows (depth=-1): dark blue background, white bold text
_SECTION_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")


def _year_code_to_fy_end(year_code: str) -> str:
    """'202526' -> '2025' (FY end year for CY), '202425' -> '2024'."""
    s = str(year_code).strip()
    if len(s) == 6:
        return f"20{s[2:4]}"  # first two digits after "20" = FY start year
    return s


def _write_master_data(ws, extractions: List[NL2Extract], existing_rows: Optional[List[list]] = None):
    """Write the Master_Data sheet — one row per company per P&L line item."""
    # Header
    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes = "A2"

    current_row = 2

    # Preserve existing rows from prior runs
    if existing_rows:
        for row_data in existing_rows:
            for col_idx, val in enumerate(row_data, 1):
                if col_idx > len(MASTER_COLUMNS):
                    break
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                col_name = MASTER_COLUMNS[col_idx - 1]
                if col_name in ("CY_Qtr", "CY_YTD", "PY_Qtr", "PY_YTD"):
                    cell.number_format = NUMBER_FORMAT
            current_row += 1

    # Write new extractions
    for extract in extractions:
        meta = get_metadata(extract.company_key)
        fy_start = _year_code_to_fy_end(extract.year) if extract.year else ""

        for pl_key in NL2_ROW_ORDER:
            depth = NL2_ROW_DEPTH.get(pl_key, 1)
            if depth == -1:
                # Section headers are display-only; they carry no data
                continue

            pl_data = extract.data.data.get(pl_key, {})

            row_values = {
                "PL_PARTICULARS":       get_pl_particulars(pl_key),
                "Grouped_PL":           get_grouped_pl(pl_key),
                "Hierarchy_Depth":      depth,
                "Company_Name":         meta["company_name"],
                "Company":              meta["sorted_company"],
                "NL":                   extract.form_type,
                "Quarter":              extract.quarter,
                "Year":                 fy_start,
                "Year_Info":            "Current Year",
                "Quarter_Info":         f"Q{extract.quarter[-1]} FY{extract.year}" if extract.year else extract.quarter,
                "Sector":               meta["sector"],
                "Industry_Competitors": meta["competitors"],
                "GI_Companies":         "GI Company",
                "CY_Qtr":               pl_data.get("cy_qtr"),
                "CY_YTD":               pl_data.get("cy_ytd"),
                "PY_Qtr":               pl_data.get("py_qtr"),
                "PY_YTD":               pl_data.get("py_ytd"),
                "Source_File":          extract.source_file,
            }

            for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
                val = row_values.get(col_name)
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if col_name in ("CY_Qtr", "CY_YTD", "PY_Qtr", "PY_YTD"):
                    cell.number_format = NUMBER_FORMAT
                # Bold + fill for summary rows (depth=0)
                if depth == 0:
                    cell.font = _BOLD_FONT
                    cell.fill = _SUMMARY_FILL

            current_row += 1


def _write_verification_sheet(ws, extract: NL2Extract):
    """Write per-company verification sheet (P&L rows x 4 period columns)."""
    ws.cell(row=1, column=1, value=f"VERIFICATION SHEET: {extract.company_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        f"Quarter: {extract.quarter} | Year: {extract.year} | Source: {extract.source_file}"
    ))

    _write_table_grid(ws, extract, start_row=4)


def _write_table_grid(ws, extract: NL2Extract, start_row: int):
    """Write the P&L grid: rows = line items, cols = 4 period columns."""
    ws.cell(row=start_row, column=1, value="NL-2 Profit and Loss Account").font = Font(bold=True)

    header_row = start_row + 1
    headers = ["Particulars", "CY Qtr", "CY YTD", "PY Qtr", "PY YTD"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_ALIGN

    for i, pl_key in enumerate(NL2_ROW_ORDER):
        ws_row = header_row + 1 + i
        pl_data = extract.data.data.get(pl_key, {})
        depth = NL2_ROW_DEPTH.get(pl_key, 1)
        display = NL2_ROW_DISPLAY_NAMES.get(pl_key, pl_key)

        label_cell = ws.cell(row=ws_row, column=1, value=display)

        if depth == -1:
            # Section header row: dark fill, white bold, spans all columns
            label_cell.font = _SECTION_FONT
            label_cell.fill = _SECTION_FILL
            for ci in range(2, 6):
                cell = ws.cell(row=ws_row, column=ci, value=None)
                cell.fill = _SECTION_FILL
            continue

        if depth == 0:
            label_cell.font = _BOLD_FONT

        for ci, period in enumerate(("cy_qtr", "cy_ytd", "py_qtr", "py_ytd"), 2):
            val = pl_data.get(period)
            cell = ws.cell(row=ws_row, column=ci, value=val)
            cell.number_format = NUMBER_FORMAT
            if depth == 0:
                cell.font = _BOLD_FONT


def _write_meta_sheet(ws, extractions: List[NL2Extract], stats: Dict[str, Any]):
    companies = sorted(set(e.company_name for e in extractions))
    quarters = sorted(set(f"{e.quarter}_{e.year}" for e in extractions))
    data = [
        ["Key", "Value"],
        ["extraction_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["extractor_version", EXTRACTOR_VERSION],
        ["files_processed", stats.get("files_processed", 0)],
        ["files_succeeded", stats.get("files_succeeded", 0)],
        ["files_failed", stats.get("files_failed", 0)],
        ["companies", ", ".join(companies)],
        ["quarters", ", ".join(quarters)],
    ]
    for r_idx, row in enumerate(data, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
            else:
                cell.fill = _META_FILL


def _sheet_name_for(extract: NL2Extract) -> str:
    name = f"{company_key_to_pascal(extract.company_key)}_{extract.quarter}_{extract.year}"
    return name[:31]


def save_workbook(extractions: List[NL2Extract], output_path: str,
                  stats: Optional[Dict[str, Any]] = None):
    if stats is None:
        stats = {}

    output_file = Path(output_path)
    existing_rows = []

    if output_file.exists():
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(output_path)
        new_files = {e.source_file for e in extractions}

        if "Master_Data" in wb.sheetnames:
            ws_old = wb["Master_Data"]
            headers = [cell.value for cell in ws_old[1]]
            if headers[:len(MASTER_COLUMNS)] == MASTER_COLUMNS:
                try:
                    sf_idx = headers.index("Source_File")
                except ValueError:
                    sf_idx = None
                if sf_idx is not None:
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        if row[sf_idx] is None:
                            continue
                        if row[sf_idx] not in new_files:
                            existing_rows.append(list(row))
            else:
                logger.warning(
                    "Existing Master_Data has different column layout -- "
                    "discarding old rows and regenerating."
                )
            del wb["Master_Data"]

        for extract in extractions:
            sn = _sheet_name_for(extract)
            if sn in wb.sheetnames:
                del wb[sn]

        if "_meta" in wb.sheetnames:
            del wb["_meta"]
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # 1. Master_Data
    ws_master = wb.create_sheet("Master_Data", 0)
    _write_master_data(ws_master, extractions, existing_rows=existing_rows)

    # 2. Verification sheets
    for extract in extractions:
        ws = wb.create_sheet(title=_sheet_name_for(extract))
        _write_verification_sheet(ws, extract)

    # 3. _meta
    ws_meta = wb.create_sheet(title="_meta")
    _write_meta_sheet(ws_meta, extractions, stats)

    wb.save(output_path)
    logger.info(f"Excel workbook saved to {output_path}")


def write_validation_summary_sheet(report_path: str, master_path: str):
    import pandas as pd
    df = pd.read_csv(report_path)
    summary = df.pivot_table(
        index=["company", "quarter", "year"],
        columns="status",
        aggfunc="size",
        fill_value=0,
    ).reset_index()
    for col in ["PASS", "WARN", "FAIL", "SKIP"]:
        if col not in summary.columns:
            summary[col] = 0
    summary["Total_Checks"] = summary[["PASS", "WARN", "FAIL"]].sum(axis=1)
    summary = summary.rename(columns={"company": "Company", "quarter": "Quarter", "year": "Year"})
    cols = ["Company", "Quarter", "Year", "Total_Checks", "PASS", "WARN", "FAIL"]
    summary = summary[cols]
    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        summary.to_excel(writer, sheet_name="Validation_Summary", index=False)


def write_validation_detail_sheet(report_path: str, master_path: str):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(report_path)
    cols_map = {
        "company": "Company", "quarter": "Quarter", "year": "Year",
        "pl_key": "PL_Key", "period": "Period", "check_name": "Check_Name",
        "status": "Status", "expected": "Expected", "actual": "Actual",
        "delta": "Delta", "note": "Note",
    }
    detail = df[df["status"].isin(["FAIL", "WARN"])].copy()
    if detail.empty:
        detail = pd.DataFrame(columns=list(cols_map.values()))
    else:
        detail = detail.rename(columns=cols_map)[list(cols_map.values())]
        detail = detail.sort_values(by="Status").reset_index(drop=True)

    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        detail.to_excel(writer, sheet_name="Validation_Detail", index=False)

    wb = load_workbook(master_path)
    ws = wb["Validation_Detail"]
    red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    status_col = list(cols_map.values()).index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col).value
        fill = red_fill if status_val == "FAIL" else yellow_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(master_path)
    logger.info(f"Validation_Detail sheet written to {master_path}")
